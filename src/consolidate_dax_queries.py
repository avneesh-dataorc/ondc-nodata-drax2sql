#!/usr/bin/env python3
"""
Script to consolidate multi-line DAX queries from Excel file into single cells.
Creates a new 'Single Line Measures' column with consolidated DAX queries.
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys

def analyze_excel_structure(file_path, sheet_name="Retail Order Dashboard Measures"):
    """
    Analyze the Excel file structure to understand merged cells and data layout.
    """
    print(f"Analyzing Excel file: {file_path}")
    print(f"Target sheet: {sheet_name}")
    
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    print(f"Sheet name: {ws.title}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")
    
    # Check for merged cells
    merged_ranges = list(ws.merged_cells.ranges)
    print(f"Number of merged cell ranges: {len(merged_ranges)}")
    
    # Show first few rows to understand structure
    print("\nFirst 10 rows of data:")
    for row in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"Row {row}: {row_data}")
    
    return wb, ws

def consolidate_dax_queries(file_path, output_path, sheet_name="Retail Order Dashboard Measures"):
    """
    Consolidate multi-line DAX queries into single cells.
    """
    print(f"Processing Excel file: {file_path}")
    print(f"Target sheet: {sheet_name}")
    
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    # Create a new workbook for output
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Retail Order Dashboard Measures"
    
    # Copy headers (no new column needed)
    headers = ["Table_name", "Measures Name", "Measures"]
    for col, header in enumerate(headers, 1):
        output_ws.cell(row=1, column=col, value=header)
    
    # Process data rows
    current_table_name = ""
    current_measure_name = ""
    current_dax_parts = []
    output_row = 2
    
    for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        table_name = ws.cell(row=row, column=1).value
        measure_name = ws.cell(row=row, column=2).value
        measure_content = ws.cell(row=row, column=3).value
        
        # Convert to string and handle None values
        table_name = str(table_name) if table_name is not None else ""
        measure_name = str(measure_name) if measure_name is not None else ""
        measure_content = str(measure_content) if measure_content is not None else ""
        
        # Check if this is a new measure (has both table_name and measure_name)
        if table_name and measure_name:
            # Save previous measure if exists
            if current_table_name and current_measure_name:
                # Consolidate DAX parts
                consolidated_dax = "\n".join(current_dax_parts)
                
                # Write to output
                output_ws.cell(row=output_row, column=1, value=current_table_name)
                output_ws.cell(row=output_row, column=2, value=current_measure_name)
                output_ws.cell(row=output_row, column=3, value=consolidated_dax)
                
                output_row += 1
            
            # Start new measure
            current_table_name = table_name
            current_measure_name = measure_name
            current_dax_parts = [measure_content] if measure_content else []
            
        else:
            # This is a continuation line (empty table_name and measure_name)
            if measure_content:
                current_dax_parts.append(measure_content)
    
    # Don't forget the last measure
    if current_table_name and current_measure_name:
        consolidated_dax = "\n".join(current_dax_parts)
        
        output_ws.cell(row=output_row, column=1, value=current_table_name)
        output_ws.cell(row=output_row, column=2, value=current_measure_name)
        output_ws.cell(row=output_row, column=3, value=consolidated_dax)
    
    # Save the output file
    output_wb.save(output_path)
    print(f"Consolidated data saved to: {output_path}")
    
    # Show summary
    total_measures = output_row - 1
    print(f"Total measures processed: {total_measures}")
    
    return output_path

def main():
    input_file = "/home/dataorc/workspace_ondc/nodata-pipline/drax/Dax query - Retail_Logistics.xlsx"
    output_file = "/home/dataorc/workspace_ondc/nodata-pipline/drax/Dax query - Retail_Logistics_consolidated.xlsx"
    
    try:
        # Analyze the structure first
        print("=== ANALYZING EXCEL STRUCTURE ===")
        wb, ws = analyze_excel_structure(input_file, "Retail Order Dashboard Measures")
        
        print("\n=== CONSOLIDATING DAX QUERIES ===")
        result_file = consolidate_dax_queries(input_file, output_file, "Retail Order Dashboard Measures")
        
        print(f"\n✅ Success! Consolidated file created: {result_file}")
        
        # Show a preview of the consolidated data
        print("\n=== PREVIEW OF CONSOLIDATED DATA ===")
        df = pd.read_excel(result_file)
        print(f"Columns: {list(df.columns)}")
        print(f"Total rows: {len(df)}")
        
        # Show first few measures
        print("\nFirst 5 measures:")
        for i in range(min(5, len(df))):
            print(f"\nMeasure {i+1}:")
            print(f"  Table: {df.iloc[i]['Table_name']}")
            print(f"  Name: {df.iloc[i]['Measures Name']}")
            print(f"  DAX (first 100 chars): {str(df.iloc[i]['Measures'])[:100]}...")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
